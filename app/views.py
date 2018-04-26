from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.views import View
from django.db.models import Q
from django.core import serializers
from openpyxl import load_workbook
from datetime import datetime

from .models import Project

# Create your views here.


class HomeView(View):
    template_name = 'home/index.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        search_type = request.POST.get('searchType')

        if search_type == 'sponsor':
            search = request.POST.get('search')
            if search:
                projects = Project.objects.filter(sponsor__contains=search, active='Y').order_by('deadline')
            else:
                projects = Project.objects.filter(active='Y').order_by('deadline')
        elif search_type == 'anything':
            anything = request.POST.get('anything')
            split = ''
            if anything.find(' + ') > -1:
                split = ' + '
            elif anything.find(' or ') > -1:
                split = ' or '
            search_queries = anything.split(split)
            sponsor = search_queries[0]
            date = search_queries[1]
            date = datetime.strptime(date, '%m/%d/%Y').strftime('%Y-%m-%d')

            if split == ' + ':
                projects = Project.objects.filter(sponsor__contains=sponsor, deadline=date, active='Y') \
                    .order_by('deadline')
            elif split == ' or ':
                projects = Project.objects.filter(Q(sponsor__contains=sponsor) | Q(deadline=date), active='Y') \
                    .order_by('deadline')
        elif search_type == 'deadline':
            deadline_from = request.POST.get('deadline_from')
            deadline_to = request.POST.get('deadline_to')
            if deadline_from and deadline_to:
                deadline_from = datetime.strptime(deadline_from, '%m/%d/%Y').strftime('%Y-%m-%d')
                deadline_to = datetime.strptime(deadline_to, '%m/%d/%Y').strftime('%Y-%m-%d')
                projects = Project.objects.filter(deadline__range=[deadline_from, deadline_to], active='Y').order_by(
                    'deadline')
            elif deadline_from:
                deadline_from = datetime.strptime(deadline_from, '%m/%d/%Y').strftime('%Y-%m-%d')
                projects = Project.objects.filter(deadline__gte=deadline_from, active='Y').order_by('deadline')
            elif deadline_to:
                deadline_to = datetime.strptime(deadline_to, '%m/%d/%Y').strftime('%Y-%m-%d')
                projects = Project.objects.filter(deadline__lte=deadline_to, active='Y').order_by('deadline')
            else:
                projects = Project.objects.all().order_by('deadline')
        else:
            min_amount = request.POST.get('min_amount')
            max_amount = request.POST.get('max_amount')
            min_amount = min_amount.replace(',', '')
            max_amount = max_amount.replace(',', '')
            if not min_amount:
                min_amount = 0
            if max_amount:
                projects = Project.objects.filter(active='Y', hide=False, amount__gte=min_amount, amount__lte=max_amount).order_by('deadline')
            else:
                projects = Project.objects.filter(active='Y', hide=False, amount__gte=min_amount).order_by('deadline')
        data = list(projects.values())
        return JsonResponse(data, safe=False)


class ImportDataView(View):
    def post(self, request, *args, **kwargs):
        print('import data')
        file = request.FILES['file']
        wb = load_workbook(filename=file)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(row_offset=1):
            sponsor = row[4].value
            if sponsor is None:
                continue
            title = row[6].value
            link = row[7].value
            amount = row[9].value
            if not isinstance(amount, int):
                amount = ''
            deadline = row[13].value
            if isinstance(deadline, str):
                deadline = None
            synopsis = row[17].value
            sponsor_deadline = row[41].value
            if isinstance(sponsor_deadline, str):
                sponsor_deadline = None
            active = row[42].value
            _type = row[43].value
            limited = row[44].value
            awards = row[45].value
            project = Project.objects.create(sponsor=sponsor, title=title, link=link, amount=amount, deadline=deadline,
                                             synopsis=synopsis, sponsor_deadline=sponsor_deadline, active=active,
                                             type=_type, limited=limited,
                                             awards=awards)
            project.save()

        return redirect('home')


class DeleteProjectView(View):
    def get(self, request, *args, **kwargs):
        project_id = request.GET.get('id')
        Project.objects.filter(id=project_id).delete()
        return redirect('home')


class ProjectView(View):
    template_name = 'project/update.html'

    def get(self, request, *args, **kwargs):
        project = Project.objects.get(id=kwargs['project_id'])
        return render(request, self.template_name, {'project': project})

    def post(self, request, *args, **kwargs):
        project = Project.objects.get(id=kwargs['project_id'])
        project.sponsor = request.POST.get('sponsor')
        project.title = request.POST.get('title')
        project.link = request.POST.get('link')
        project.amount = request.POST.get('amount')
        project.synopsis = request.POST.get('synopsis')
        project.active = request.POST.get('active')
        project.type = request.POST.get('type')
        project.limited = request.POST.get('limited')
        project.awards = request.POST.get('awards')
        project.limited = request.POST.get('limited')
        deadline = request.POST.get('deadline')
        if deadline:
            project.deadline = datetime.strptime(deadline, '%m/%d/%Y')
        else:
            project.deadline = None
        sponsor_deadline = request.POST.get('sponsor_deadline')
        if deadline:
            project.sponsor_deadline = datetime.strptime(sponsor_deadline, '%m/%d/%Y')
        else:
            project.deadline = None
        project.comment = request.POST.get('comment')
        hide = request.POST.get('hide')
        if hide == 'Y':
            project.hide = True
        else:
            project.hide = False
        project.save()

        return render(request, self.template_name, {'project': project})

